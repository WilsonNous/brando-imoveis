from flask import Flask, render_template, request, redirect, jsonify, send_file, session
from datetime import datetime
from io import StringIO
import csv
import config
from models import db, Imovel, Lead, Servico, ImovelFoto
import io
import logging
from sqlalchemy.pool import QueuePool
import os

# ============================================================
# IMPORTS PARA EXCEL / PARSING
# ============================================================

import re
from decimal import Decimal, InvalidOperation

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:
    openpyxl = None
    Workbook = None
    DataValidation = None


# ============================================================
# INICIALIZAÇÃO FLASK + BANCO
# ============================================================

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = config.SQLALCHEMY_DATABASE_URI
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = config.SQLALCHEMY_TRACK_MODIFICATIONS
app.secret_key = config.SECRET_KEY

# 🔐 Senha do painel admin (Render)
ADMIN_PASSWORD = config.ADMIN_PASSWORD


# ============================================================
# LOGIN DO ADMIN
# ============================================================

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        senha = request.form.get("senha")

        if senha == ADMIN_PASSWORD:
            session["admin_auth"] = True
            return redirect("/admin")

        return render_template("admin_login.html", erro="Senha incorreta.")

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_auth", None)
    return redirect("/admin/login")


def require_admin():
    if not session.get("admin_auth"):
        return redirect("/admin/login")


# ============================================================
# FILTRO DE FORMATAÇÃO MONETÁRIA (pt-BR)
# ============================================================

@app.template_filter('brl')
def format_brl(value):
    try:
        value = float(value)
        return f"R$ {value:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


# ============================================================
# ENGINE OPTIONS — MYSQL HOSTGATOR
# ============================================================

app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 280,
    "pool_size": 5,
    "max_overflow": 10,
    "poolclass": QueuePool,
    "connect_args": {"connect_timeout": 10},
}

db.init_app(app)


# ============================================================
# CONFIGURAÇÃO DE IMAGENS (apenas validação de tipo)
# ============================================================

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
app.logger.info("🚀 Brando Imóveis iniciado com pool seguro.")


# ============================================================
# HELPERS IMPORT/EXPORT (VALOR BRL + CSV DIALECT + STATUS)
# ============================================================

def parse_valor_brl(raw) -> float:
    """
    Converte valores em formatos comuns (Excel/CSV) para float (reais):
      - 1200000
      - 1200000.00
      - 1.200.000,00
      - R$ 1.200.000,00
      - "R$ 260.000,00"
    """
    if raw is None:
        return 0.0

    s = str(raw).strip()
    if not s:
        return 0.0

    # remove moeda e espaços
    s = s.replace("R$", "").replace("r$", "").strip()
    s = s.replace("\u00a0", " ").replace(" ", "")

    # remove qualquer coisa que não seja dígito, ponto, vírgula, sinal
    s = re.sub(r"[^0-9\-,.]", "", s)

    # se tem . e , -> padrão BR: 1.234.567,89
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    # se tem só vírgula -> vira decimal
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    # se tem só ponto -> já é decimal US/Excel ou inteiro com ponto

    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return 0.0


def sniff_csv_dialect(text: str) -> str:
    """
    Detecta delimitador (',' ou ';') usando heurística no header.
    """
    sample = "\n".join(text.splitlines()[:20])
    header = sample.splitlines()[0] if sample.splitlines() else ""
    delim = ";" if header.count(";") > header.count(",") else ","
    return delim


def normalize_status(raw, fallback="ativo") -> str:
    """
    Garante apenas 'ativo' ou 'inativo'.
    Se vier qualquer outra coisa, usa fallback.
    """
    s = (str(raw).strip().lower() if raw is not None else "")
    if s in ("ativo", "inativo"):
        return s
    return fallback or "ativo"


# ============================================================
# ROTA PARA SERVIR FOTOS (BLOB)
# ============================================================

@app.route("/foto/<int:foto_id>")
def foto_blob(foto_id):
    foto = ImovelFoto.query.get_or_404(foto_id)
    if not foto.conteudo:
        return "Imagem não encontrada.", 404

    mimetype = foto.mimetype or "image/jpeg"
    return send_file(io.BytesIO(foto.conteudo), mimetype=mimetype)


# ============================================================
# PÁGINAS PÚBLICAS
# ============================================================

@app.route('/')
def home():
    try:
        imoveis = Imovel.query.filter_by(status='ativo').all()
        return render_template('index.html', imoveis=imoveis)
    except Exception as e:
        app.logger.error(f"❌ Erro ao consultar imóveis: {e}")
        return "Erro ao conectar ao banco.", 500


@app.route('/imovel/<int:id>')
def imovel_detalhe(id):
    imovel = Imovel.query.get(id)
    if not imovel:
        return "Imóvel não encontrado.", 404

    fotos = [f"/foto/{f.id}" for f in imovel.fotos]

    # Se não tiver fotos BLOB mas tiver imagem antiga
    if not fotos and getattr(imovel, "imagem", None):
        fotos.append(imovel.imagem)

    if not fotos:
        fotos = ["https://picsum.photos/800/600?blur=1"]

    return render_template("imovel.html", imovel=imovel, fotos=fotos)


# ============================================================
# DEFINIR CAPA / REMOVER FOTO
# ============================================================

@app.route("/admin/imovel/<int:imovel_id>/set_capa/<int:foto_id>", methods=["POST"])
def set_capa(imovel_id, foto_id):
    r = require_admin()
    if r: return r

    imovel = Imovel.query.get_or_404(imovel_id)
    foto = ImovelFoto.query.get_or_404(foto_id)

    for f in imovel.fotos:
        f.is_capa = (f.id == foto.id)

    # compat com coluna antiga
    if hasattr(imovel, "imagem"):
        imovel.imagem = f"/foto/{foto.id}"

    db.session.commit()
    return redirect(f"/admin/edit/{imovel_id}")


@app.route("/admin/imovel/<int:imovel_id>/remove_foto/<int:foto_id>", methods=["POST"])
def remove_foto(imovel_id, foto_id):
    r = require_admin()
    if r: return r

    imovel = Imovel.query.get_or_404(imovel_id)
    foto = ImovelFoto.query.get_or_404(foto_id)

    was_capa = foto.is_capa
    db.session.delete(foto)
    db.session.commit()

    imovel = Imovel.query.get_or_404(imovel_id)
    if was_capa and imovel.fotos:
        imovel.fotos[0].is_capa = True
        if hasattr(imovel, "imagem"):
            imovel.imagem = f"/foto/{imovel.fotos[0].id}"
        db.session.commit()

    return redirect(f"/admin/edit/{imovel_id}")


# ============================================================
# LEADS
# ============================================================

@app.route('/lead', methods=['POST'])
def lead():
    try:
        nome = request.form['nome']
        telefone = request.form['telefone']
        mensagem = request.form.get('mensagem', '')
        imovel_id = request.form.get('imovel_id')

        novo_lead = Lead(
            nome=nome,
            telefone=telefone,
            mensagem=mensagem,
            imovel_id=imovel_id,
            data=datetime.now()
        )
        db.session.add(novo_lead)
        db.session.commit()

        msg = (
            f"Olá! Tenho interesse no imóvel código {imovel_id}"
            if imovel_id else
            "Olá! Tenho interesse em imóveis da Brando."
        )
        return redirect(f"https://wa.me/5548991054216?text={msg}")

    except Exception as e:
        app.logger.error(f"❌ Erro ao registrar lead: {e}")
        return "Erro ao enviar lead.", 500


# ============================================================
# PAINEL ADMIN — LISTA / EDITAR / SALVAR
# ============================================================

@app.route('/admin')
def admin():
    r = require_admin()
    if r: return r

    q = request.args.get('q', '')
    base = Imovel.query

    if q:
        like = f"%{q}%"
        base = base.filter(
            Imovel.codigo.ilike(like) |
            Imovel.tipo.ilike(like) |
            Imovel.bairro.ilike(like) |
            Imovel.descricao.ilike(like)
        )

    imoveis = base.order_by(Imovel.id.desc()).all()
    return render_template('admin.html', imoveis=imoveis, imovel=None)


@app.route('/admin/edit/<int:id>')
def admin_edit(id):
    r = require_admin()
    if r: return r

    imovel = Imovel.query.get(id)
    imoveis = Imovel.query.order_by(Imovel.id.desc()).all()
    return render_template('admin.html', imoveis=imoveis, imovel=imovel)


@app.route('/admin/delete/<int:id>')
def admin_delete(id):
    r = require_admin()
    if r: return r

    item = Imovel.query.get(id)
    if item:
        db.session.delete(item)
        db.session.commit()
    return redirect('/admin')


# ============================================================
# SALVAR IMÓVEL + UPLOAD MÚLTIPLO (BLOB)
# ============================================================

@app.route('/admin/save', methods=['POST'])
def admin_save():
    r = require_admin()
    if r: return r

    form = request.form
    iid = form.get('id')

    obj = Imovel.query.get(int(iid)) if iid else Imovel()
    if not iid:
        db.session.add(obj)

    obj.codigo = form.get('codigo')
    obj.tipo = form.get('tipo')
    obj.bairro = form.get('bairro')
    obj.descricao = form.get('descricao')
    obj.status = normalize_status(form.get('status'), fallback=obj.status or "ativo")

    try:
        obj.valor = float((form.get('valor') or "0").replace(',', '.'))
    except Exception:
        obj.valor = 0.0

    files = request.files.getlist('imagens')
    ja_tem_fotos = len(obj.fotos) > 0

    for file in files:
        if file and file.filename and allowed_file(file.filename):
            conteudo = file.read()
            if not conteudo:
                continue

            foto = ImovelFoto(
                imovel=obj,
                conteudo=conteudo,
                mimetype=file.mimetype or "image/jpeg"
            )

            if not ja_tem_fotos and not any(f.is_capa for f in obj.fotos):
                foto.is_capa = True

            db.session.add(foto)
            ja_tem_fotos = True

    db.session.commit()

    if obj.fotos and not any(f.is_capa for f in obj.fotos):
        obj.fotos[0].is_capa = True
        db.session.commit()

    return redirect('/admin')


# ============================================================
# EXPORTAÇÃO / IMPORTAÇÃO (Excel XLSX principal + CSV fallback)
# ============================================================

@app.route('/admin/modelo.xlsx')
def admin_modelo_xlsx():
    """
    Baixa o MODELO OFICIAL BRANDO para preenchimento e importação sem erro.
    """
    r = require_admin()
    if r: return r

    if openpyxl is None or Workbook is None:
        return "openpyxl não está instalado no servidor. Adicione no requirements.txt", 500

    wb = Workbook()
    ws = wb.active
    ws.title = "imoveis"

    headers = ["codigo", "tipo", "valor", "bairro", "descricao", "status"]
    ws.append(headers)

    # Exemplo (ajuda a preencher certo)
    ws.append([
        "A001",
        "apartamento",
        260000.00,
        "Ingleses",
        "Exemplo: 2 quartos, 1 suíte, sacada com churrasqueira...",
        "ativo"
    ])

    ws.freeze_panes = "A2"

    # Larguras
    col_widths = {"A": 12, "B": 18, "C": 14, "D": 18, "E": 60, "F": 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Formata coluna valor como moeda no Excel (visual)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=2000):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    # Validação dropdown status
    if DataValidation is not None:
        dv_status = DataValidation(type="list", formula1='"ativo,inativo"', allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add("F2:F2000")

    # Aba de instruções
    ws2 = wb.create_sheet("LEIA-ME")
    ws2.append(["MODELO OFICIAL — Como preencher"])
    ws2.append(["1) Valor: use número puro. Ex: 260000 ou 260000,00 (não digite 'R$')."])
    ws2.append(["2) Status deve ser: ativo ou inativo (use o dropdown)."])
    ws2.append(["3) Código é obrigatório e deve ser único (A001, C003, T010...)."])
    ws2.append(["4) Mantenha os nomes das colunas exatamente como no header."])
    ws2.column_dimensions["A"].width = 90

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="modelo_imoveis_brando.xlsx"
    )


@app.route('/admin/export.xlsx')
def admin_export_xlsx():
    r = require_admin()
    if r: return r

    if openpyxl is None or Workbook is None:
        return "openpyxl não está instalado no servidor. Adicione no requirements.txt", 500

    wb = Workbook()
    ws = wb.active
    ws.title = "imoveis"

    ws.append(["codigo", "tipo", "valor", "bairro", "descricao", "status"])

    for i in Imovel.query.order_by(Imovel.id).all():
        ws.append([
            i.codigo,
            i.tipo,
            float(i.valor or 0),
            i.bairro,
            i.descricao or "",
            normalize_status(i.status, "ativo"),
        ])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="imoveis.xlsx"
    )


@app.route('/admin/export')  # mantém CSV (fallback)
def admin_export_csv():
    r = require_admin()
    if r: return r

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['codigo', 'tipo', 'valor', 'bairro', 'descricao', 'status'])

    for i in Imovel.query.order_by(Imovel.id).all():
        writer.writerow([
            i.codigo,
            i.tipo,
            float(i.valor or 0),  # número puro no CSV (sem R$)
            i.bairro,
            (i.descricao or '').replace("\n", " "),
            normalize_status(i.status, "ativo")
        ])

    output = si.getvalue().encode('utf-8')
    return send_file(
        io.BytesIO(output),
        mimetype='text/csv',
        as_attachment=True,
        download_name='imoveis.csv'
    )


@app.route('/admin/import', methods=['POST'])
def admin_import():
    """
    Importa tanto XLSX quanto CSV.
    - Principal: XLSX (Excel)
    - Fallback: CSV com delimitador detectado e encoding robusto
    """
    r = require_admin()
    if r:
        return r

    file = request.files.get('planilha')
    if not file or not file.filename:
        return redirect('/admin')

    filename = (file.filename or "").lower()
    count = 0

    expected_headers = ["codigo", "tipo", "valor", "bairro", "descricao", "status"]

    # ------------------------
    # XLSX (principal)
    # ------------------------
    if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        if openpyxl is None:
            return "openpyxl não está instalado no servidor. Adicione no requirements.txt", 500

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        headers = []
        for cell in ws[1]:
            headers.append((str(cell.value).strip().lower() if cell.value is not None else ""))

        missing = [h for h in expected_headers if h not in headers]
        if missing:
            return (
                f"Planilha inválida. Faltando colunas: {', '.join(missing)}. "
                f"Use o modelo oficial (/admin/modelo.xlsx)."
            ), 400

        def col_idx(name: str):
            return headers.index(name)

        idx_codigo = col_idx("codigo")
        idx_tipo = col_idx("tipo")
        idx_valor = col_idx("valor")
        idx_bairro = col_idx("bairro")
        idx_desc = col_idx("descricao")
        idx_status = col_idx("status")
        idx_imagem = headers.index("imagem") if "imagem" in headers else None  # opcional

        for row in ws.iter_rows(min_row=2, values_only=True):
            # pula linha totalmente vazia
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            codigo = (str(row[idx_codigo]).strip() if row[idx_codigo] is not None else "")
            if not codigo:
                continue

            obj = Imovel.query.filter_by(codigo=codigo).first()
            if not obj:
                obj = Imovel(codigo=codigo)
                db.session.add(obj)

            obj.tipo = (str(row[idx_tipo]).strip() if row[idx_tipo] is not None else (obj.tipo or ""))
            obj.bairro = (str(row[idx_bairro]).strip() if row[idx_bairro] is not None else (obj.bairro or ""))
            obj.descricao = (str(row[idx_desc]).strip() if row[idx_desc] is not None else (obj.descricao or ""))

            obj.status = normalize_status(row[idx_status], fallback=obj.status or "ativo")
            obj.valor = parse_valor_brl(row[idx_valor])

            # imagem opcional (se existir no modelo)
            if idx_imagem is not None and hasattr(obj, "imagem"):
                img = (str(row[idx_imagem]).strip() if row[idx_imagem] is not None else "")
                if img:
                    obj.imagem = img

            count += 1

        db.session.commit()
        app.logger.info(f"✅ Importados/atualizados via XLSX: {count}")
        return redirect('/admin')

    # ------------------------
    # CSV (fallback)
    # ------------------------
    raw = file.stream.read()
    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')

    delim = sniff_csv_dialect(text)
    stream = io.StringIO(text)
    reader = csv.DictReader(stream, delimiter=delim)

    csv_headers = [h.strip().lower() for h in (reader.fieldnames or [])]
    missing = [h for h in expected_headers if h not in csv_headers]
    if missing:
        return (
            f"CSV inválido. Faltando colunas: {', '.join(missing)}. "
            f"Use o modelo oficial (/admin/modelo.xlsx)."
        ), 400

    for row in reader:
        codigo = (row.get('codigo') or '').strip()
        if not codigo:
            continue

        obj = Imovel.query.filter_by(codigo=codigo).first()
        if not obj:
            obj = Imovel(codigo=codigo)
            db.session.add(obj)

        obj.tipo = (row.get('tipo') or obj.tipo or '').strip()
        obj.bairro = (row.get('bairro') or obj.bairro or '').strip()
        obj.descricao = (row.get('descricao') or obj.descricao or '').strip()
        obj.status = normalize_status(row.get('status'), fallback=obj.status or "ativo")

        v = (row.get('valor') or '').strip()
        if v:
            obj.valor = parse_valor_brl(v)

        if hasattr(obj, "imagem"):
            img = (row.get('imagem') or '').strip()
            if img:
                obj.imagem = img

        count += 1

    db.session.commit()
    app.logger.info(f"✅ Importados/atualizados via CSV: {count} (delim='{delim}')")
    return redirect('/admin')


# ============================================================
# SERVIÇOS PÚBLICO E ADMIN
# ============================================================

@app.route("/servicos", methods=["GET", "POST"])
def servicos():
    sucesso, erro = False, None

    if request.method == "POST":
        try:
            nome = request.form.get("nome_cliente")
            telefone = request.form.get("telefone")
            imovel_id = request.form.get("imovel_id")
            tipo_servico = request.form.get("tipo_servico")
            descricao = request.form.get("descricao")

            imovel_id = int(imovel_id) if imovel_id and imovel_id.isdigit() else None

            novo = Servico(
                nome_cliente=nome,
                telefone=telefone,
                imovel_id=imovel_id,
                tipo_servico=tipo_servico,
                descricao=descricao,
                data_solicitacao=datetime.now(),
                status="pendente",
            )

            db.session.add(novo)
            db.session.commit()
            sucesso = True

        except Exception as e:
            db.session.rollback()
            erro = str(e)

    imoveis = Imovel.query.filter_by(status="ativo").order_by(Imovel.id.desc()).all()
    return render_template("servicos.html", sucesso=sucesso, erro=erro, imoveis=imoveis)


@app.route('/admin/servicos')
def admin_servicos():
    r = require_admin()
    if r: return r

    servicos = (
        db.session.query(Servico)
        .outerjoin(Imovel)
        .options(db.contains_eager(Servico.imovel))
        .order_by(Servico.data_solicitacao.desc())
        .all()
    )
    return render_template('admin_servicos.html', servicos=servicos)


@app.route('/admin/servicos/update/<int:id>', methods=['POST'])
def update_servico(id):
    r = require_admin()
    if r: return r

    s = Servico.query.get(id)
    if not s:
        return redirect('/admin/servicos')

    s.status = request.form.get('status', s.status)
    data_ag = request.form.get('data_agendamento')
    if data_ag:
        s.data_agendamento = datetime.strptime(data_ag, "%Y-%m-%d")
    s.responsavel = request.form.get('responsavel')
    s.materiais = request.form.get('materiais')

    custo = request.form.get('custo', '0').replace(',', '.')
    try:
        s.custo = float(custo)
    except Exception:
        s.custo = 0.0

    db.session.commit()
    return redirect('/admin/servicos')


# ============================================================
# PÁGINA DE TEMPORADA
# ============================================================

@app.route('/temporada')
def temporada():
    imoveis = Imovel.query.filter(Imovel.status == 'ativo').all()
    return render_template('temporada.html', imoveis=imoveis)


# ============================================================
# RUN
# ============================================================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("✅ Banco de dados ok")
    app.run(debug=True)
